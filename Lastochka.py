# Ошибка Nonetype возникала на документах 7, 8, 9, 10 из-за того, что в методе settype использовалась не сама строка, а substring. И в методе substring ыла ошибка, что в случае, если len(a)<=1 возвращается None. В данном случае, если в строке нет пробелов, len(a)=1 и выдаётся None. А должно быть при len(a)<1. На это None, возвращаемое, как substring и вводимое в функцию поиска синонимов, как аргумент, возвращало NoneType
# Попутно я внедрил две проверки на none массива. Их можно убрать при желании или проблеме


import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment
import re

bend_radius = []
og_bend_radius = []
flanec_fulfilments = []

types = []
steels = []
in_coatings = []
out_coatings = []
perehod_types = []

angles = []
diameters = []
walls = []
pipe_lengths = []
og_angles = []

flanec_types = []
pressures = []
uslov_diameters = []

# функция получает row - всю, которую нужно обслужить
def linear_list_fulfillment(row, array):
    col=2
    while sheet.cell(row=row, column=col).value is not None:
        # Получаем значение из ячейки
        value_from_cell = sheet.cell(row=row, column=col).value
        array.append(value_from_cell)
        col += 1

# функция получает row-координату ячейки, с которой нужно начать вправо вниз
def double_measure_list_fulfillment(row, array):
    i = -1
    col = 2
    # супер!
    while sheet.cell(row=row, column=2).value is not None:
        array.append([])
        i += 1
        col = 2
        while sheet.cell(row=row, column=col).value is not None:
            value_from_cell = sheet.cell(row=row, column=col).value
            array[i].append(value_from_cell)
            # Получаем значение из ячейки и вставляем в массив
            col += 1
        row += 1

# Открываем файл Excel
workbook1 = openpyxl.load_workbook("числовые, enumerate значения, типы стали.xlsx")
# Выбираем активный лист (или можно выбрать конкретный лист по имени)
sheet = workbook1.active


#чётенько
linear_list_fulfillment(2, bend_radius)
linear_list_fulfillment(3, og_bend_radius)
linear_list_fulfillment(4, flanec_fulfilments)




linear_list_fulfillment(7, angles)
linear_list_fulfillment(8, diameters)
linear_list_fulfillment(9, walls)
linear_list_fulfillment(10, pipe_lengths)
linear_list_fulfillment(11, og_angles)

linear_list_fulfillment(13, flanec_types)
linear_list_fulfillment(14, pressures)
linear_list_fulfillment(15, uslov_diameters)

#print(bend_radius)
#print(og_bend_radius)
#print(flanec_fulfilments)

#print(angles)
#print(diameters)
#print(walls)
#print(pipe_lengths)
#print(og_angles)

#print(flanec_types)
#print(pressures)
#print(uslov_diameters)
# чётенько


double_measure_list_fulfillment(18, perehod_types)
double_measure_list_fulfillment(22, steels)
#print(perehod_types)
#print(steels)
# работает!


# Открываем второй файл Excel
workbook2 = openpyxl.load_workbook("типы покрытий.xlsx")
# Выбираем активный лист (или можно выбрать конкретный лист по имени)
sheet = workbook2.active

double_measure_list_fulfillment(1, in_coatings)
double_measure_list_fulfillment(4, out_coatings)
#print(in_coatings)
#print(out_coatings)

# Открываем третий файл Excel
workbook3 = openpyxl.load_workbook("типы фитингов.xlsx")
# Выбираем активный лист (или можно выбрать конкретный лист по имени)
sheet = workbook3.active

double_measure_list_fulfillment(1, types)
#print(types)


























class Fitting:
    def init(self, characteristics):
        self.characteristics = None

class Otvod(Fitting):
    def _init_(self):
        self.characteristics = None
        self.integers = []
        self.angle = None  # Угол (15, 30, 45, 60, 90)
        self.diameter = None  # Диаметр
        self.wall_thickness = None  # Стенка (2.5 - 32.0)
        self.bend_radius = None  # Радиус гиба (1DN или 1.5DN)
        self.steel = None  # Сталь (09Г2С, ст.20, 13ХФА, и др.)
        self.pipe_length = None  # Длина патрубка (целое число > 100, 0 или None)
        self.coating = None  # Покрытие (без покрытия, с покрытием, комбинация покрытий)
    def set_parameters(self):
        self.bend_radius = "1DN" if "1DN" in self.characteristics else ""
        if "1DN" in self.characteristics:
            self.characteristics=self.characteristics.replace("1DN", '')
        if "1,5DN" in self.characteristics:
            self.characteristics=self.characteristics.replace("1,5DN", '')

        self.steel = find_steel(self.characteristics)
        self.coating = find_coatings(self.characteristics)

        
        integers = extract_integers_from_string(self.characteristics)

        angle = find_angle(integers)
        self.angle = angle
        
        diameter, wall = find_diameter_and_wall_thickness(integers)
        self.diameter, self.wall_thickness = diameter, wall

        length = find_length(integers)
        self.pipe_length = length

    def to_string(self):
        return f"Отвод {self.angle}° {self.diameter}х{self.wall_thickness} {self.bend_radius} {self.steel} {self.pipe_length} {self.coating}\n\n"

class Perehod(Fitting):
    def _init_(self):
        self.characteristics = None
        self.integers = []
        self.type = None  # Тип (К - концентрический, Э - эксцентрический)
        self.diameter = None  # Диаметр1
        self.wall_thickness = None  # Стенка1 (2.5 - 32.0)
        self.diameter2 = None  # Диаметр2 (< Диаметр1)
        self.wall_thickness2 = None  # Стенка2 (<= Стенка1)
        self.steel = None  # Сталь (09Г2С, ст.20, 13ХФА, и др.)
        self.pipe_length = None  # Длина патрубка (целое число > 100, 0 или None)
        self.coating = None  # Покрытие (без покрытия, с покрытием, комбинация покрытий)
    def set_parameters(self):

        self.type = find_perehod_type(self.characteristics)
        self.steel = find_steel(self.characteristics)
        self.coating = find_coatings(self.characteristics)
        
        integers = extract_integers_from_string(self.characteristics)
        
        diameter, wall = find_diameter_and_wall_thickness(integers)
        self.diameter, self.wall_thickness = diameter, wall

        
        #if(diameter in integers):
         #   integers.remove(diameter)
        #if(wall in integers):
         #   integers.remove(wall)
        diameter2, wall2 = find_diameter_and_wall_thickness(integers)
        self.diameter2, self.wall_thickness2 = diameter2, wall2

        length = find_length(integers)
        self.pipe_length = length
        

    def to_string(self):
        return f"Переход {self.type} {self.diameter}х{self.wall_thickness}-{self.diameter2}х{self.wall_thickness2} {self.steel} {self.pipe_length} {self.coating}\n\n"


class Troinik(Fitting):
    def _init_(self):
        self.characteristics = None
        self.integers = []
        self.diameter = None  # Диаметр магистрального патрубка
        self.wall_thickness = None  # Стенка магистрального патрубка (2.5 - 32.0)
        self.diameter_otv = None  # Диаметр отвода (для равнопроходных тройников)
        self.wall_thickness_otv = None  # Стенка отвода (для равнопроходных тройников) #он херово будет искать вторую толщину. Пофиксить потом
        self.steel = None  # Сталь (09Г2С, ст.20, 13ХФА, и др.)
        self.pipe_length = None  # Длина патрубка (целое число > 100, 0 или None)
        self.coating = None  # Покрытие (без покрытия, с покрытием, комбинация покрытий)
    def set_parameters(self):
        self.steel = find_steel(self.characteristics)
        self.coating = find_coatings(self.characteristics)

        
        integers = extract_integers_from_string(self.characteristics)

        
        diameter, wall = find_diameter_and_wall_thickness(integers)
        self.diameter, self.wall_thickness = diameter, wall

        
        #if(diameter in integers):
         #   integers.remove(diameter)
        #if(wall in integers):
         #   integers.remove(wall)
        diameter2, wall2 = find_diameter_and_wall_thickness(integers)
        self.diameter_otv, self.wall_thickness_otv = diameter2, wall2

        length = find_length(integers)
        self.pipe_length = length
        
    def to_string(self):
        if(self.diameter_otv != self.diameter and self.diameter_otv != None and self.diameter != None):
            return f"Тройник {self.diameter}х{self.wall_thickness}-{self.diameter_otv}х{self.wall_thickness_otv} {self.steel} {self.pipe_length}\n\n"
        else:
            return f"Тройник {self.diameter}х{self.wall_thickness} {self.steel} {self.pipe_length}\n\n"


class Zaglushka(Fitting):
    def _init_(self):
        self.characteristics = None
        self.integers = []
        self.detail = None
        self.diameter = None  # Диаметр
        self.wall_thickness = None  # Стенка (2.5 - 32.0)
        self.steel = None  # Сталь (09Г2С, ст.20, 13ХФА, и др.)
        self.pipe_length = None  # Длина патрубка (целое число > 100, 0 или None)
        self.coating = None  # Покрытие (без покрытия, с покрытием, комбинация покрытий)
    def set_parameters(self):
        self.steel = find_steel(self.characteristics)
        self.coating = find_coatings(self.characteristics)
        
        integers = extract_integers_from_string(self.characteristics)
        
        diameter, wall = find_diameter_and_wall_thickness(integers)
        self.diameter, self.wall_thickness = diameter, wall
        if(self.diameter!=None):
            self.detail = "Заглушка" if self.diameter <= 426 else "Днище"
        else:
            self.detail = "Заглушка"


        length = find_length(integers)
        self.pipe_length = length
        
        
    def to_string(self):
        return f"{self.detail} {self.diameter}х{self.wall_thickness} {self.steel} {self.pipe_length} {self.coating}\n\n"


class Flanec(Fitting):
    def _init_(self):
        self.characteristics = None
        self.integers = []
        self.diameter_conditional = None  # Диаметр условный
        self.pressure = None  # Давление
        self.flange_type = None  # Тип (01 - плоский, 11 - воротниковый)
        self.execution = None  # Исполнение (B, E, F, J)
        self.steel = None  # Сталь (09Г2С, ст.20, 13ХФА, и др.)
        self.pipe_diameter = None  # Диаметр трубы
        self.wall_thickness = None  # Стенка (2.5 - 32.0)
        self.pipe_length = None  # Длина патрубка (целое число > 100)
        self.coating = None  # Покрытие (без покрытия, с покрытием, комбинация покрытий)
    def set_parameters(self):

        self.execution = find_flanec_fulfilments(self.characteristics)

        self.steel = find_steel(self.characteristics)
        self.coating = find_coatings(self.characteristics)

        integers = extract_integers_from_string(self.characteristics)

        cond_diameter, pressure, diameter, wall = find_uslov_diameter_pressure_diam_and_wall(integers)
        self.diameter_conditional, self.pressure, self.pipe_diameter, self.wall_thickness = cond_diameter, pressure, diameter, wall


        length = find_length(integers)
        self.pipe_length = length

        
        self.flange_type = find_flanec_type(self.characteristics, integers)

        
    def to_string(self):
        return f"Фланец {self.diameter_conditional}-{self.pressure}-{self.flange_type}-{self.execution} {self.steel} {self.pipe_diameter}х{self.wall_thickness} {self.pipe_length} {self.coating}\n\n"


class OtvodOG(Fitting):
    def _init_(self):
        self.characteristics = None
        self.integers = []
        self.angle = None  # Угол (1 - 90 градусов)
        print(self.integers)
        self.diameter = None  # Диаметр
        self.wall_thickness = None  # Стенка (4 - 32.0 мм)
        self.steel = None  # Сталь (09Г2С, ст.20, 13ХФА, и др.)
        self.bend_radius = None  # Радиус гиба (3DN - 25DN)
        self.construction_length = None  # Строительная длина (цифра, кратная 50, от 250 до 10000)
        self.coating = None  # Покрытие (без покрытия, с покрытием, комбинация покрытий)
    def set_parameters(self):

        self.bend_radius = find_og_bend_radius(self.characteristics)

        self.steel = find_steel(self.characteristics)
        self.coating = find_coatings(self.characteristics)

        integers = extract_integers_from_string(self.characteristics)
        
        angle_og=find_og_angle(integers)
        self.angle = angle_og
        diameter, wall = find_diameter_and_wall_thickness(integers)
        self.diameter, self.wall_thickness = diameter, wall

        length = find_length(integers)
        self.construction_length = length
        

        
    def to_string(self):
        return f"Отвод ОГ {self.angle}° {self.diameter}х{self.wall_thickness} {self.steel} {self.bend_radius} {self.construction_length} {self.coating}\n\n"


class Default_fitting(Fitting):
    def _init_(self, characteristics):
        super().__init__(characteristics)
    def set_parameters(self):
        self.characteristics="Не удалось определить номенклатуру фитинга: "+self.characteristics
    def to_string(self):
        return self.characteristics
        pass























    

def find_first_element_in_string(input_string, elements_to_find):
    for element in elements_to_find:
        if str(element) in input_string:
            input_string=input_string.replace(str(element), '')
            return element
    return None  # Если ничего не найдено

def find_first_synonym_in_string(input_string, synonym_array):

    returning = None
    gold_array = []
    for synonyms in synonym_array:
        if synonyms is None:
            return None  # Добавьте обработку случая, когда synonyms равно None
    
        if synonyms[0] in input_string:
            if synonyms[0] is None:
                return None
            
            input_string=input_string.replace(synonyms[0], '')
            if (returning == None):
                returning = synonyms[0]

    for synonyms in synonym_array:
        if synonyms is None:
            return None  # Добавьте обработку случая, когда synonyms равно None
    
        for words in synonyms:
            if words in input_string:
                input_string=input_string.replace(str(words), '')
                if(returning == None):
                    returning = synonyms[0]
    return returning
    # программа проходится по всем словам, чтобы удалить все повторы и вернуть последний элемент

def extract_integers_from_string(input_string):
    if "(" in input_string:
        input_string=input_string.replace("(", ' ')
    if ")" in input_string:
        input_string=input_string.replace(")", ' ')
    # Используем регулярное выражение для поиска всех целых чисел в строке
    integers = re.findall(r'\d+', input_string)
    # Преобразуем найденные строки в целые числа
    integers = [int(num) for num in integers]
    return integers

def find_first_integer_in_array(parameters_from_string, parameters_array):
    for integers in parameters_from_string:
        if integers in parameters_array:
            parameters_from_string.remove(integers)
            return integers
    return None
#мы не можем работать сразу со строкой, поскольку нет гарантии, что в ней нет другого числа, содержащего наше в записи
#придётся держать массив числовых значений

def extract_substring(input_string):
    # Разделяем строку по пробелу
    parts = input_string.split(' ')
    
    # Если строка содержит хотя бы один пробел и после него есть хотя бы 4 символа
    if len(parts) > 1 and len(parts[1]) >= 4:
        # Формируем новую строку, объединяя первую часть до пробела, пробел и первые 4 символа после пробела
        result_string = parts[0] + ' ' + parts[1][:4]
        return result_string
    
    if len(parts) > 1 and len(parts[1]) < 4:
        # Формируем новую строку, объединяя первую часть до пробела, пробел и первые 4 символа после пробела
        result_string = parts[0] + ' ' + parts[1]
        return result_string
    if len(parts) < 1:
        return None
    # Если не удовлетворяет условию, возвращаем исходную строку
    return input_string







def find_bend_radius(input_string):
    bend_radii=find_first_element_in_string(input_string, bend_radius)
    return bend_radii

def find_og_bend_radius(input_string):
    og_bend_radii=find_first_element_in_string(input_string, og_bend_radius)
    return og_bend_radii

def find_flanec_fulfilments(input_string):
    flanec_fulfilment=find_first_element_in_string(input_string, flanec_fulfilments)
    return flanec_fulfilment






def settype(input_string):
    in_string = extract_substring(input_string)
    a = find_first_synonym_in_string(in_string, types)
    if(a=="Отвод"):
        b=Otvod()
        return b
    if(a=="Переход"):
        b=Perehod()
        return b
    if(a=="Тройник"):
        b=Troinik()
        return b
    if(a=="Заглушка" or a=="Днище"):
        b=Zaglushka()
        return b
    if(a=="Фланец"):
        b=Flanec()
        return b
    if(a=="Отвод ОГ"):
        b=OtvodOG()
        return b
    else:
        b=Default_fitting()
        return b
    
def find_steel(input_string):
    steel = find_first_synonym_in_string(input_string, steels)
    return steel

def find_coatings(input_string):
    in_coating = find_first_synonym_in_string(input_string, in_coatings)
    out_coating = find_first_synonym_in_string(input_string, out_coatings)
    returning = ""
    if(in_coating!=None and out_coating!=None):
        returning = in_coating+"/"+out_coating
    elif(in_coating != None):
        returning = in_coating
    elif(out_coating != None):
        returning = out_coating
    return returning

def find_perehod_type(input_string):
    perehod_type=find_first_synonym_in_string(input_string, perehod_types)
    return perehod_type












def find_diameter_and_wall_thickness(input_array):
    diameter = find_first_integer_in_array(input_array, diameters)
    if(diameter!=None):
        if (diameter < 57):
            thickness = find_first_integer_in_array(input_array, walls[:22])
        elif(diameter >= 57 and diameter <= 107):
            thickness = find_first_integer_in_array(input_array, walls[:22])
        elif(diameter >= 108 and diameter <= 218):
            thickness = find_first_integer_in_array(input_array, walls[:22])
        elif(diameter >= 219 and diameter <= 425):
            thickness = find_first_integer_in_array(input_array, walls[3:24])
        elif(diameter >= 426):
            thickness = find_first_integer_in_array(input_array, walls[5:])
    else:
        thickness = find_first_integer_in_array(input_array, walls)
    return diameter, thickness

def find_length(input_array):
    length = find_first_integer_in_array(input_array, pipe_lengths)
    if length!=None:
        return "L="+str(length)
    else:   return ""

def find_angle(input_array):
    angle = find_first_integer_in_array(input_array, angles)
    return angle

def find_og_angle(input_array):
    og_angle = find_first_integer_in_array(input_array, og_angles)
    return og_angle

def find_pressure(input_array):
    pressure = find_first_integer_in_array(input_array, pressures)
    return pressure

def find_flanec_type(input_string, input_array):
    if("01" in input_string):
        input_string=input_string.replace("01", '')
        return "01"
    elif(11 in input_array):
        return 11
    else:
        return None

def find_uslov_diameter_pressure_diam_and_wall(parameters_from_string):
    i=-1
    usl_d=None
    for number in uslov_diameters:
        i+=1
        if number in parameters_from_string:
            parameters_from_string.remove(number)
            usl_d = number
            break
    pressure = find_first_integer_in_array(parameters_from_string, pressures)
    if(usl_d!=None):
        diam = diameters[i]
    else:
        diam = find_first_integer_in_array(parameters_from_string, diameters)
    if(diameter!=None):
            if (diam < 57):
                thickness = find_first_integer_in_array(parameters_from_string, walls[:22])
            elif(diam >= 57 and diameter <= 107):
                thickness = find_first_integer_in_array(parameters_from_string, walls[:22])
            elif(diam >= 108 and diameter <= 218):
                thickness = find_first_integer_in_array(parameters_from_string, walls[:22])
            elif(diam >= 219 and diameter <= 425):
                thickness = find_first_integer_in_array(parameters_from_string, walls[3:24])
            else:
                thickness = find_first_integer_in_array(parameters_from_string, walls[5:])    
    else:
        thickness = find_first_integer_in_array(parameters_from_string, walls)
    wall = thickness
    #if(diam!=None):
     #   parameters_from_string.remove(diam)
    return usl_d, pressure, diam, wall





































































# Запуск прохода по таблице
print("Введите название и полный путь к файлу с расширением xlsx")
filename=input()
# Открываем файл Excel
workbook = openpyxl.load_workbook(filename)

# Выбираем активный лист (или можно выбрать конкретный лист по имени)
sheet = workbook.active

# Начинаем цикл с A2 и итерируемся до тех пор, пока не встретим пустую ячейку
row = 2  # Начинаем с строки 2 (A2)

# Инициализируем три массива для каждой категории
arr1, arr2, arr3 = [], [], []

while sheet.cell(row=row, column=1).value is not None:
    # Получаем значение из левой ячейки (столбец A)
    value_from_left = sheet.cell(row=row, column=1).value

    integers = str(extract_integers_from_string(value_from_left))

    # Применяем функцию settype (замените на вашу функцию)
    fitting = settype(value_from_left)
    fitting.integers = integers
    fitting.characteristics = value_from_left
    fitting.set_parameters()
    # Получаем значение новой номенклатуры
    nomenclature=fitting.to_string()

    #print(extract_integers_from_string(value_from_left))
    diameter, wall = find_diameter_and_wall_thickness(extract_integers_from_string(value_from_left))
    if "Не удалось" in nomenclature:
        arr3.append([value_from_left, nomenclature])
    elif "None" in nomenclature:
        arr2.append([value_from_left, nomenclature])
    else:
        arr1.append([value_from_left, nomenclature])
    # Переходим к следующей строке
    row += 1

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.value = None

new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# Увеличьте ширину колонок (в данном случае увеличим колонки A и B в 3 раза)
new_ws.column_dimensions['A'].width = 6 * new_ws.column_dimensions['A'].width
new_ws.column_dimensions['B'].width = 6 * new_ws.column_dimensions['B'].width

# Создаем стиль для заголовка
header_style = NamedStyle(name='header_style')
header_style.font = Font(bold=True, size=33)  # Установите полужирный шрифт
header_style.alignment = Alignment(horizontal='center', vertical='center')  # Выравнивание по центру

# Создаем стиль для ячейки с переносом текста
wrap_text_style = NamedStyle(name='wrap_text_style')
wrap_text_style.alignment = Alignment(wrapText=True)


# Вставляем данные в ячейки
new_ws['A1'] = "Названия из заявки"
new_ws['B1'] = "Названия по номенклатуре"
new_ws['A2'] = "удалось отсортировать без ошибок"
new_ws['A1'].style = header_style
new_ws['B1'].style = header_style
new_ws['A2'].style = wrap_text_style

# Функция для вставки массива в ячейки
def insert_array(arr, start_row):
    for i, row in enumerate(arr):
        for j, value in enumerate(row):
            cell = new_ws.cell(row=start_row + i, column=j + 1, value=value)
            cell.style=wrap_text_style

# Определяем строки для вставки
start_row = 3

# Вставляем первый массив
insert_array(arr1, start_row)
start_row += len(arr1) + 2  # Добавляем 2 строки пропуска

# Вставляем второй массив
cell = new_ws.cell(row=start_row, column=1, value="возможны ошибки")
cell.style=wrap_text_style
insert_array(arr2, start_row + 1)
start_row += len(arr2) + 2

# Вставляем третий массив
cell = new_ws.cell(row=start_row, column=1, value="не удалось обработать")
cell.style=wrap_text_style
insert_array(arr3, start_row + 1)

# Сохраняем изменения в файл
new_wb.save(filename+" с посчитанной номенклатурой.xlsx")
print("Данные записаны в файл "+filename+" с посчитанной номенклатурой.xlsx")
# Пока функция сравнивает выданное значение с "правильным значением". Это правильно. Функция find first element in string будет и дальше выдавать правильный ответ, только потом она будет выбирать ещё и из неправильных правильный
